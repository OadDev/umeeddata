from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import bcrypt
import jwt
import secrets
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from pydantic import BaseModel, Field, EmailStr
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ.get("JWT_SECRET", secrets.token_hex(32))

# Create the main app
app = FastAPI(title="Umeed Now Foundation Finance Dashboard")
api_router = APIRouter(prefix="/api")

# ============ MODELS ============
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "user"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: datetime

class CampaignCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    commission_percentage: float = Field(ge=0, le=100)
    company_share: float = Field(ge=0, le=100)
    dev_share: float = Field(ge=0, le=100)
    himanshu_share: float = Field(ge=0, le=100)
    denim_share: float = Field(ge=0, le=100)
    status: str = "active"

class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    commission_percentage: Optional[float] = None
    company_share: Optional[float] = None
    dev_share: Optional[float] = None
    himanshu_share: Optional[float] = None
    denim_share: Optional[float] = None
    status: Optional[str] = None

class DailyEntryCreate(BaseModel):
    campaign_id: str
    date: str  # YYYY-MM-DD format
    ad_spend: float = Field(ge=0)
    website_collection: float = Field(ge=0)
    qr_collection: float = Field(ge=0)

class DailyEntryUpdate(BaseModel):
    ad_spend: Optional[float] = None
    website_collection: Optional[float] = None
    qr_collection: Optional[float] = None

class MonthlySettlementCreate(BaseModel):
    campaign_id: str
    report_month: str  # YYYY-MM format
    ad_account_charges: float = Field(ge=0)
    miscellaneous_expenses: float = Field(ge=0)
    notes: Optional[str] = ""

class MonthlySettlementUpdate(BaseModel):
    ad_account_charges: Optional[float] = None
    miscellaneous_expenses: Optional[float] = None
    notes: Optional[str] = None

class SettingsUpdate(BaseModel):
    gst_percentage: Optional[float] = None
    gateway_percentage: Optional[float] = None

# ============ HELPERS ============
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def get_jwt_secret() -> str:
    return JWT_SECRET

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["id"] = str(user["_id"])
        del user["_id"]
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

async def get_settings():
    settings = await db.settings.find_one({"_id": "global"}, {"_id": 0})
    if not settings:
        settings = {"gst_percentage": 18.0, "gateway_percentage": 2.6}
    return settings

def calculate_daily_values(entry: dict, settings: dict, commission_percentage: float) -> dict:
    ad_spend = entry.get("ad_spend", 0)
    website_collection = entry.get("website_collection", 0)
    qr_collection = entry.get("qr_collection", 0)
    gst = settings.get("gst_percentage", 18.0)
    gateway = settings.get("gateway_percentage", 2.6)
    
    ad_spend_with_gst = round(ad_spend + (ad_spend * gst / 100), 2)
    gateway_charge = round(website_collection * gateway / 100, 2)
    total_revenue = round(website_collection + qr_collection - gateway_charge, 2)
    net_profit = round(total_revenue - ad_spend_with_gst, 2)
    platform_commission = round(net_profit * commission_percentage / 100, 2) if net_profit > 0 else 0
    
    return {
        "ad_spend_with_gst": ad_spend_with_gst,
        "gateway_charge": gateway_charge,
        "total_revenue": total_revenue,
        "net_profit": net_profit,
        "platform_commission": platform_commission
    }

# ============ AUTH ROUTES ============
@api_router.post("/auth/register")
async def register(user_data: UserCreate, response: Response):
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed = hash_password(user_data.password)
    user_doc = {
        "email": email,
        "password_hash": hashed,
        "name": user_data.name,
        "role": user_data.role,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {"id": user_id, "email": email, "name": user_data.name, "role": user_data.role}

@api_router.post("/auth/login")
async def login(user_data: UserLogin, response: Response, request: Request):
    email = user_data.email.lower()
    ip = request.client.host
    identifier = f"{ip}:{email}"
    
    # Check brute force
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        lockout_time = attempt.get("last_attempt", datetime.now(timezone.utc))
        if datetime.now(timezone.utc) - lockout_time < timedelta(minutes=15):
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again in 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})
    
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(user_data.password, user["password_hash"]):
        # Increment failed attempts
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc)}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Clear failed attempts on success
    await db.login_attempts.delete_one({"identifier": identifier})
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    
    return {"id": user_id, "email": email, "name": user["name"], "role": user["role"]}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return user

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(data: PasswordChange, user: dict = Depends(get_current_user)):
    # Get user with password hash
    db_user = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password
    if not verify_password(data.current_password, db_user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Validate new password
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    
    # Update password
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {"password_hash": new_hash}}
    )
    
    return {"message": "Password changed successfully"}

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Refresh token missing")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        
        access_token = create_access_token(str(user["_id"]), user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
        return {"message": "Token refreshed"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ============ USER MANAGEMENT (ADMIN) ============
@api_router.get("/users")
async def get_users(admin: dict = Depends(require_admin)):
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    for user in users:
        user["id"] = str(user["_id"])
        del user["_id"]
    return users

@api_router.post("/users")
async def create_user(user_data: UserCreate, admin: dict = Depends(require_admin)):
    email = user_data.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed = hash_password(user_data.password)
    user_doc = {
        "email": email,
        "password_hash": hashed,
        "name": user_data.name,
        "role": user_data.role,
        "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(user_doc)
    return {"id": str(result.inserted_id), "email": email, "name": user_data.name, "role": user_data.role}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

# ============ SETTINGS (ADMIN) ============
@api_router.get("/settings")
async def get_settings_route(user: dict = Depends(get_current_user)):
    settings = await get_settings()
    return settings

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate, admin: dict = Depends(require_admin)):
    update_data = {}
    if data.gst_percentage is not None:
        update_data["gst_percentage"] = data.gst_percentage
    if data.gateway_percentage is not None:
        update_data["gateway_percentage"] = data.gateway_percentage
    
    if update_data:
        await db.settings.update_one(
            {"_id": "global"},
            {"$set": update_data},
            upsert=True
        )
    return await get_settings()

# ============ CAMPAIGNS ============
@api_router.get("/campaigns")
async def get_campaigns(user: dict = Depends(get_current_user)):
    campaigns = await db.campaigns.find({}, {"_id": 0}).to_list(1000)
    return campaigns

@api_router.get("/campaigns/{campaign_id}")
async def get_campaign(campaign_id: str, user: dict = Depends(get_current_user)):
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return campaign

@api_router.post("/campaigns")
async def create_campaign(data: CampaignCreate, user: dict = Depends(get_current_user)):
    # Validate commission splits
    total_split = data.company_share + data.dev_share + data.himanshu_share + data.denim_share
    if abs(total_split - data.commission_percentage) > 0.01:
        raise HTTPException(status_code=400, detail=f"Commission splits ({total_split}%) must equal total commission ({data.commission_percentage}%)")
    
    campaign_id = str(ObjectId())
    campaign_doc = {
        "id": campaign_id,
        "name": data.name,
        "description": data.description,
        "commission_percentage": data.commission_percentage,
        "company_share": data.company_share,
        "dev_share": data.dev_share,
        "himanshu_share": data.himanshu_share,
        "denim_share": data.denim_share,
        "status": data.status,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc)
    }
    await db.campaigns.insert_one(campaign_doc)
    campaign_doc.pop("_id", None)
    return campaign_doc

@api_router.put("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, data: CampaignUpdate, admin: dict = Depends(require_admin)):
    campaign = await db.campaigns.find_one({"id": campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    update_data = {}
    for field, value in data.model_dump(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    # Validate commission splits if any are being updated
    if any(k in update_data for k in ["commission_percentage", "company_share", "dev_share", "himanshu_share", "denim_share"]):
        commission = update_data.get("commission_percentage", campaign["commission_percentage"])
        company = update_data.get("company_share", campaign["company_share"])
        dev = update_data.get("dev_share", campaign["dev_share"])
        himanshu = update_data.get("himanshu_share", campaign["himanshu_share"])
        denim = update_data.get("denim_share", campaign["denim_share"])
        total_split = company + dev + himanshu + denim
        if abs(total_split - commission) > 0.01:
            raise HTTPException(status_code=400, detail=f"Commission splits ({total_split}%) must equal total commission ({commission}%)")
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.campaigns.update_one({"id": campaign_id}, {"$set": update_data})
    
    return await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})

@api_router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, admin: dict = Depends(require_admin)):
    result = await db.campaigns.delete_one({"id": campaign_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Campaign not found")
    # Also delete related entries
    await db.daily_entries.delete_many({"campaign_id": campaign_id})
    await db.monthly_settlements.delete_many({"campaign_id": campaign_id})
    return {"message": "Campaign deleted"}

# ============ DAILY ENTRIES ============
@api_router.get("/daily-entries")
async def get_daily_entries(
    campaign_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if campaign_id:
        query["campaign_id"] = campaign_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    entries = await db.daily_entries.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    return entries

@api_router.post("/daily-entries")
async def create_daily_entry(data: DailyEntryCreate, user: dict = Depends(get_current_user)):
    # Check for duplicate
    existing = await db.daily_entries.find_one({"campaign_id": data.campaign_id, "date": data.date})
    if existing:
        raise HTTPException(status_code=400, detail="Entry for this date already exists")
    
    # Get campaign for commission percentage
    campaign = await db.campaigns.find_one({"id": data.campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    settings = await get_settings()
    calculated = calculate_daily_values(data.model_dump(), settings, campaign["commission_percentage"])
    
    entry_id = str(ObjectId())
    entry_doc = {
        "id": entry_id,
        "campaign_id": data.campaign_id,
        "date": data.date,
        "ad_spend": data.ad_spend,
        "website_collection": data.website_collection,
        "qr_collection": data.qr_collection,
        **calculated,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc)
    }
    await db.daily_entries.insert_one(entry_doc)
    entry_doc.pop("_id", None)
    return entry_doc

@api_router.put("/daily-entries/{entry_id}")
async def update_daily_entry(entry_id: str, data: DailyEntryUpdate, admin: dict = Depends(require_admin)):
    entry = await db.daily_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    update_data = {}
    for field, value in data.model_dump(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    if update_data:
        # Recalculate values
        campaign = await db.campaigns.find_one({"id": entry["campaign_id"]})
        settings = await get_settings()
        merged = {**entry, **update_data}
        calculated = calculate_daily_values(merged, settings, campaign["commission_percentage"])
        update_data.update(calculated)
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.daily_entries.update_one({"id": entry_id}, {"$set": update_data})
    
    return await db.daily_entries.find_one({"id": entry_id}, {"_id": 0})

@api_router.delete("/daily-entries/{entry_id}")
async def delete_daily_entry(entry_id: str, admin: dict = Depends(require_admin)):
    result = await db.daily_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}

# CSV Upload for bulk import
class CSVUploadResponse(BaseModel):
    success: int
    failed: int
    errors: List[str]

@api_router.post("/daily-entries/upload-csv/{campaign_id}")
async def upload_csv_entries(campaign_id: str, file_content: str = Query(...), user: dict = Depends(get_current_user)):
    """
    Upload CSV data for bulk import.
    CSV format: date,ad_spend,website_collection,qr_collection
    Date format: YYYY-MM-DD or DD/MM/YYYY or DD-MM-YYYY
    """
    import csv
    from io import StringIO
    
    campaign = await db.campaigns.find_one({"id": campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    settings = await get_settings()
    
    success_count = 0
    failed_count = 0
    errors = []
    
    try:
        reader = csv.DictReader(StringIO(file_content))
        for row_num, row in enumerate(reader, start=2):
            try:
                # Parse date - support multiple formats
                date_str = row.get('date', '').strip()
                if not date_str:
                    errors.append(f"Row {row_num}: Missing date")
                    failed_count += 1
                    continue
                
                # Try different date formats
                parsed_date = None
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                
                if not parsed_date:
                    errors.append(f"Row {row_num}: Invalid date format '{date_str}'")
                    failed_count += 1
                    continue
                
                formatted_date = parsed_date.strftime('%Y-%m-%d')
                
                # Parse numeric values
                try:
                    ad_spend = float(row.get('ad_spend', 0) or 0)
                    website_collection = float(row.get('website_collection', 0) or 0)
                    qr_collection = float(row.get('qr_collection', 0) or 0)
                except ValueError as e:
                    errors.append(f"Row {row_num}: Invalid number format - {str(e)}")
                    failed_count += 1
                    continue
                
                # Check for duplicate
                existing = await db.daily_entries.find_one({"campaign_id": campaign_id, "date": formatted_date})
                if existing:
                    errors.append(f"Row {row_num}: Entry for {formatted_date} already exists")
                    failed_count += 1
                    continue
                
                # Calculate values
                entry_data = {
                    "ad_spend": ad_spend,
                    "website_collection": website_collection,
                    "qr_collection": qr_collection
                }
                calculated = calculate_daily_values(entry_data, settings, campaign["commission_percentage"])
                
                # Insert entry
                entry_id = str(ObjectId())
                entry_doc = {
                    "id": entry_id,
                    "campaign_id": campaign_id,
                    "date": formatted_date,
                    **entry_data,
                    **calculated,
                    "created_by": user["id"],
                    "created_at": datetime.now(timezone.utc)
                }
                await db.daily_entries.insert_one(entry_doc)
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                failed_count += 1
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid CSV format: {str(e)}")
    
    return {"success": success_count, "failed": failed_count, "errors": errors[:20]}

# ============ MONTHLY SETTLEMENTS ============
@api_router.get("/monthly-settlements")
async def get_monthly_settlements(
    campaign_id: Optional[str] = None,
    report_month: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    if campaign_id:
        query["campaign_id"] = campaign_id
    if report_month:
        query["report_month"] = report_month
    
    settlements = await db.monthly_settlements.find(query, {"_id": 0}).to_list(1000)
    return settlements

@api_router.post("/monthly-settlements")
async def create_monthly_settlement(data: MonthlySettlementCreate, admin: dict = Depends(require_admin)):
    # Check for duplicate
    existing = await db.monthly_settlements.find_one({"campaign_id": data.campaign_id, "report_month": data.report_month})
    if existing:
        raise HTTPException(status_code=400, detail="Settlement for this month already exists")
    
    campaign = await db.campaigns.find_one({"id": data.campaign_id})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    settlement_id = str(ObjectId())
    settlement_doc = {
        "id": settlement_id,
        "campaign_id": data.campaign_id,
        "report_month": data.report_month,
        "ad_account_charges": data.ad_account_charges,
        "miscellaneous_expenses": data.miscellaneous_expenses,
        "notes": data.notes,
        "entered_by": admin["id"],
        "entered_on": datetime.now(timezone.utc)
    }
    await db.monthly_settlements.insert_one(settlement_doc)
    settlement_doc.pop("_id", None)
    return settlement_doc

@api_router.put("/monthly-settlements/{settlement_id}")
async def update_monthly_settlement(settlement_id: str, data: MonthlySettlementUpdate, admin: dict = Depends(require_admin)):
    settlement = await db.monthly_settlements.find_one({"id": settlement_id})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    update_data = {}
    for field, value in data.model_dump(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.monthly_settlements.update_one({"id": settlement_id}, {"$set": update_data})
    
    return await db.monthly_settlements.find_one({"id": settlement_id}, {"_id": 0})

@api_router.delete("/monthly-settlements/{settlement_id}")
async def delete_monthly_settlement(settlement_id: str, admin: dict = Depends(require_admin)):
    result = await db.monthly_settlements.delete_one({"id": settlement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Settlement not found")
    return {"message": "Settlement deleted"}

# ============ DASHBOARD ============
@api_router.get("/dashboard")
async def get_dashboard(user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    yesterday = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    
    # This month range
    this_month_start = now.replace(day=1).strftime("%Y-%m-%d")
    this_month_end = today
    
    # Last month range
    last_month_end = (now.replace(day=1) - timedelta(days=1))
    last_month_start = last_month_end.replace(day=1).strftime("%Y-%m-%d")
    last_month_end = last_month_end.strftime("%Y-%m-%d")
    
    settings = await get_settings()
    
    # Today's global stats
    today_entries = await db.daily_entries.find({"date": today}, {"_id": 0}).to_list(1000)
    
    # Yesterday's stats
    yesterday_entries = await db.daily_entries.find({"date": yesterday}, {"_id": 0}).to_list(1000)
    
    # This month's stats
    this_month_entries = await db.daily_entries.find({
        "date": {"$gte": this_month_start, "$lte": this_month_end}
    }, {"_id": 0}).to_list(10000)
    
    # Last month's stats
    last_month_entries = await db.daily_entries.find({
        "date": {"$gte": last_month_start, "$lte": last_month_end}
    }, {"_id": 0}).to_list(10000)
    
    def calculate_stats(entries):
        return {
            "total_profit": round(sum(e.get("net_profit", 0) for e in entries), 2),
            "platform_profit": round(sum(e.get("platform_commission", 0) for e in entries), 2),
            "total_revenue": round(sum(e.get("total_revenue", 0) for e in entries), 2),
            "total_ad_spend": round(sum(e.get("ad_spend", 0) for e in entries), 2),
            "website_collection": round(sum(e.get("website_collection", 0) for e in entries), 2),
            "qr_collection": round(sum(e.get("qr_collection", 0) for e in entries), 2)
        }
    
    today_stats = calculate_stats(today_entries)
    yesterday_stats = calculate_stats(yesterday_entries)
    this_month_stats = calculate_stats(this_month_entries)
    last_month_stats = calculate_stats(last_month_entries)
    
    # Calculate percentage changes
    def calc_change(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return round(((current - previous) / abs(previous)) * 100, 1)
    
    today_vs_yesterday = {
        "profit_change": calc_change(today_stats["total_profit"], yesterday_stats["total_profit"]),
        "revenue_change": calc_change(today_stats["total_revenue"], yesterday_stats["total_revenue"]),
        "ad_spend_change": calc_change(today_stats["total_ad_spend"], yesterday_stats["total_ad_spend"])
    }
    
    this_month_vs_last = {
        "profit_change": calc_change(this_month_stats["total_profit"], last_month_stats["total_profit"]),
        "revenue_change": calc_change(this_month_stats["total_revenue"], last_month_stats["total_revenue"]),
        "ad_spend_change": calc_change(this_month_stats["total_ad_spend"], last_month_stats["total_ad_spend"])
    }
    
    # Campaign-wise stats
    campaigns = await db.campaigns.find({"status": "active"}, {"_id": 0}).to_list(1000)
    campaign_stats = []
    for camp in campaigns:
        camp_entries = [e for e in today_entries if e["campaign_id"] == camp["id"]]
        campaign_stats.append({
            "id": camp["id"],
            "name": camp["name"],
            "profit": sum(e.get("net_profit", 0) for e in camp_entries),
            "platform_profit": sum(e.get("platform_commission", 0) for e in camp_entries),
            "revenue": sum(e.get("total_revenue", 0) for e in camp_entries)
        })
    
    # Get last 30 days trend data
    thirty_days_ago = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    trend_entries = await db.daily_entries.find({"date": {"$gte": thirty_days_ago}}, {"_id": 0}).to_list(10000)
    
    # Group by date for trends
    trend_data = {}
    for entry in trend_entries:
        date = entry["date"]
        if date not in trend_data:
            trend_data[date] = {"date": date, "revenue": 0, "profit": 0, "ad_spend": 0}
        trend_data[date]["revenue"] += entry.get("total_revenue", 0)
        trend_data[date]["profit"] += entry.get("net_profit", 0)
        trend_data[date]["ad_spend"] += entry.get("ad_spend", 0)
    
    trend_list = sorted(trend_data.values(), key=lambda x: x["date"])
    
    return {
        "today": today_stats,
        "yesterday": yesterday_stats,
        "this_month": this_month_stats,
        "last_month": last_month_stats,
        "today_vs_yesterday": today_vs_yesterday,
        "this_month_vs_last": this_month_vs_last,
        "campaign_stats": campaign_stats,
        "trend_data": trend_list,
        "settings": settings
    }

# ============ REPORTS ============
@api_router.get("/reports")
async def get_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    campaign_id: Optional[str] = None,
    preset: Optional[str] = None,  # today, yesterday, day_before, this_month
    user: dict = Depends(get_current_user)
):
    today = datetime.now(timezone.utc)
    
    # Handle presets
    if preset == "today":
        start_date = end_date = today.strftime("%Y-%m-%d")
    elif preset == "yesterday":
        yesterday = today - timedelta(days=1)
        start_date = end_date = yesterday.strftime("%Y-%m-%d")
    elif preset == "day_before":
        day_before = today - timedelta(days=2)
        start_date = end_date = day_before.strftime("%Y-%m-%d")
    elif preset == "this_month":
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    
    query = {}
    if campaign_id:
        query["campaign_id"] = campaign_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    entries = await db.daily_entries.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    
    # Enrich with campaign names
    campaigns = {c["id"]: c for c in await db.campaigns.find({}, {"_id": 0}).to_list(1000)}
    for entry in entries:
        camp = campaigns.get(entry["campaign_id"], {})
        entry["campaign_name"] = camp.get("name", "Unknown")
    
    # Summary
    summary = {
        "total_ad_spend": round(sum(e.get("ad_spend", 0) for e in entries), 2),
        "total_ad_spend_with_gst": round(sum(e.get("ad_spend_with_gst", 0) for e in entries), 2),
        "total_website_collection": round(sum(e.get("website_collection", 0) for e in entries), 2),
        "total_qr_collection": round(sum(e.get("qr_collection", 0) for e in entries), 2),
        "total_gateway_charge": round(sum(e.get("gateway_charge", 0) for e in entries), 2),
        "total_revenue": round(sum(e.get("total_revenue", 0) for e in entries), 2),
        "total_net_profit": round(sum(e.get("net_profit", 0) for e in entries), 2),
        "total_platform_commission": round(sum(e.get("platform_commission", 0) for e in entries), 2),
        "entry_count": len(entries)
    }
    
    return {"entries": entries, "summary": summary}

# ============ MONTHLY REPORTS ============
@api_router.get("/monthly-reports")
async def get_monthly_reports(
    report_month: Optional[str] = None,
    campaign_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    campaigns = await db.campaigns.find({}, {"_id": 0}).to_list(1000)
    campaign_map = {c["id"]: c for c in campaigns}
    
    # Filter campaigns
    if campaign_id:
        campaigns = [c for c in campaigns if c["id"] == campaign_id]
    
    # Get all months with data
    pipeline = [
        {"$group": {"_id": {"$substr": ["$date", 0, 7]}}},
        {"$sort": {"_id": -1}}
    ]
    months = [m["_id"] for m in await db.daily_entries.aggregate(pipeline).to_list(100)]
    
    if report_month and report_month in months:
        months = [report_month]
    
    reports = []
    for camp in campaigns:
        for month in months:
            # Get daily entries for this campaign and month
            month_start = f"{month}-01"
            # Calculate month end
            year, mon = int(month[:4]), int(month[5:7])
            if mon == 12:
                month_end = f"{year+1}-01-01"
            else:
                month_end = f"{year}-{mon+1:02d}-01"
            
            entries = await db.daily_entries.find({
                "campaign_id": camp["id"],
                "date": {"$gte": month_start, "$lt": month_end}
            }, {"_id": 0}).to_list(1000)
            
            if not entries:
                continue
            
            funds_raised = sum(e.get("website_collection", 0) + e.get("qr_collection", 0) for e in entries)
            ad_cost_with_gst = sum(e.get("ad_spend_with_gst", 0) for e in entries)
            gateway_charge = sum(e.get("gateway_charge", 0) for e in entries)
            net_profit = sum(e.get("net_profit", 0) for e in entries)
            commission = sum(e.get("platform_commission", 0) for e in entries)
            
            # Get settlement
            settlement = await db.monthly_settlements.find_one({
                "campaign_id": camp["id"],
                "report_month": month
            }, {"_id": 0})
            
            ad_account_charges = settlement.get("ad_account_charges", 0) if settlement else 0
            misc_expenses = settlement.get("miscellaneous_expenses", 0) if settlement else 0
            
            funds_to_give = net_profit - commission - ad_account_charges - misc_expenses
            
            # Calculate stakeholder earnings
            stakeholder_earnings = {
                "company": round(commission * (camp["company_share"] / camp["commission_percentage"]), 2) if camp["commission_percentage"] > 0 else 0,
                "dev": round(commission * (camp["dev_share"] / camp["commission_percentage"]), 2) if camp["commission_percentage"] > 0 else 0,
                "himanshu": round(commission * (camp["himanshu_share"] / camp["commission_percentage"]), 2) if camp["commission_percentage"] > 0 else 0,
                "denim": round(commission * (camp["denim_share"] / camp["commission_percentage"]), 2) if camp["commission_percentage"] > 0 else 0
            }
            
            reports.append({
                "campaign_id": camp["id"],
                "campaign_name": camp["name"],
                "month": month,
                "funds_raised": round(funds_raised, 2),
                "ad_cost_with_gst": round(ad_cost_with_gst, 2),
                "gateway_charge": round(gateway_charge, 2),
                "net_profit": round(net_profit, 2),
                "commission": round(commission, 2),
                "ad_account_charges": round(ad_account_charges, 2),
                "miscellaneous_expenses": round(misc_expenses, 2),
                "funds_to_give": round(funds_to_give, 2),
                "stakeholder_earnings": stakeholder_earnings,
                "has_settlement": settlement is not None
            })
    
    return reports

# ============ STAKEHOLDER EARNINGS ============
@api_router.get("/stakeholder-earnings")
async def get_stakeholder_earnings(
    campaign_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    # Get all monthly reports
    reports = await get_monthly_reports(campaign_id=campaign_id, user=user)
    
    # Aggregate by stakeholder
    totals = {"company": 0, "dev": 0, "himanshu": 0, "denim": 0}
    by_month = {}
    
    for report in reports:
        month = report["month"]
        if month not in by_month:
            by_month[month] = {"company": 0, "dev": 0, "himanshu": 0, "denim": 0}
        
        for stakeholder in ["company", "dev", "himanshu", "denim"]:
            amount = report["stakeholder_earnings"].get(stakeholder, 0)
            totals[stakeholder] += amount
            by_month[month][stakeholder] += amount
    
    return {
        "totals": {k: round(v, 2) for k, v in totals.items()},
        "by_month": {k: {sk: round(sv, 2) for sk, sv in v.items()} for k, v in by_month.items()}
    }

# ============ PDF GENERATION ============
def format_date_short(date_str):
    """Format date as '1 Feb 26' for PDF table"""
    from datetime import datetime
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    day = date_obj.day
    return f"{day} {date_obj.strftime('%b %y')}"

def format_month_name(month_str):
    """Format month as 'March, 2026'"""
    from datetime import datetime
    date_obj = datetime.strptime(f"{month_str}-01", "%Y-%m-%d")
    return date_obj.strftime("%B, %Y")

def format_rupees_compact(value):
    """Format currency compactly for PDF table"""
    return f"{value:,.0f}"

def format_rupees_full(value):
    """Format currency with Rs. prefix for summary"""
    return f"Rs. {value:,.2f}"

@api_router.get("/generate-pdf/{campaign_id}/{report_month}")
async def generate_pdf(campaign_id: str, report_month: str, user: dict = Depends(get_current_user)):
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    settings = await get_settings()
    
    # Get entries for this month
    month_start = f"{report_month}-01"
    year, mon = int(report_month[:4]), int(report_month[5:7])
    if mon == 12:
        month_end = f"{year+1}-01-01"
    else:
        month_end = f"{year}-{mon+1:02d}-01"
    
    entries = await db.daily_entries.find({
        "campaign_id": campaign_id,
        "date": {"$gte": month_start, "$lt": month_end}
    }, {"_id": 0}).sort("date", 1).to_list(1000)
    
    # Get settlement
    settlement = await db.monthly_settlements.find_one({
        "campaign_id": campaign_id,
        "report_month": report_month
    }, {"_id": 0})
    
    # Format month name for display
    month_display = format_month_name(report_month)
    
    # Create PDF with wider margins
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=15, leftMargin=15, topMargin=25, bottomMargin=25)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=15, textColor=colors.HexColor('#6AAF35'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=8)
    
    elements = []
    
    # Header
    elements.append(Paragraph(f"Umeed Now Foundation - Monthly Report", title_style))
    elements.append(Paragraph(f"Campaign: {campaign['name']}", subtitle_style))
    elements.append(Paragraph(f"Month: {month_display} | Commission: {campaign['commission_percentage']}% | GST: {settings['gst_percentage']}% | Gateway: {settings['gateway_percentage']}%", subtitle_style))
    elements.append(Spacer(1, 10))
    
    # Table data - include commission percentage in headers
    comm_pct = campaign['commission_percentage']
    after_comm_pct = 100 - comm_pct
    table_data = [["Date", "Ad Spend", "Ad+GST", "Website", "QR", "Gateway", "Revenue", "Profit", f"Comm({comm_pct}%)", f"After({after_comm_pct}%)"]]
    
    totals = {"ad_spend": 0, "ad_spend_with_gst": 0, "website": 0, "qr": 0, "gateway": 0, "revenue": 0, "profit": 0, "commission": 0, "after_commission": 0}
    
    for entry in entries:
        after_comm = entry['net_profit'] - entry['platform_commission']
        table_data.append([
            format_date_short(entry["date"]),
            format_rupees_compact(entry['ad_spend']),
            format_rupees_compact(entry['ad_spend_with_gst']),
            format_rupees_compact(entry['website_collection']),
            format_rupees_compact(entry['qr_collection']),
            format_rupees_compact(entry['gateway_charge']),
            format_rupees_compact(entry['total_revenue']),
            format_rupees_compact(entry['net_profit']),
            format_rupees_compact(entry['platform_commission']),
            format_rupees_compact(after_comm)
        ])
        totals["ad_spend"] += entry["ad_spend"]
        totals["ad_spend_with_gst"] += entry["ad_spend_with_gst"]
        totals["website"] += entry["website_collection"]
        totals["qr"] += entry["qr_collection"]
        totals["gateway"] += entry["gateway_charge"]
        totals["revenue"] += entry["total_revenue"]
        totals["profit"] += entry["net_profit"]
        totals["commission"] += entry["platform_commission"]
        totals["after_commission"] += after_comm
    
    # Totals row
    table_data.append([
        "TOTAL",
        format_rupees_compact(totals['ad_spend']),
        format_rupees_compact(totals['ad_spend_with_gst']),
        format_rupees_compact(totals['website']),
        format_rupees_compact(totals['qr']),
        format_rupees_compact(totals['gateway']),
        format_rupees_compact(totals['revenue']),
        format_rupees_compact(totals['profit']),
        format_rupees_compact(totals['commission']),
        format_rupees_compact(totals['after_commission'])
    ])
    
    # Create table with explicit column widths to fit landscape A4 (842 points width - 30 margins = 812)
    col_widths = [55, 70, 70, 75, 60, 60, 75, 75, 75, 75]  # Total ~690 points
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6AAF35')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Date column left aligned
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F5F5')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary section
    ad_account = settlement.get("ad_account_charges", 0) if settlement else 0
    misc = settlement.get("miscellaneous_expenses", 0) if settlement else 0
    funds_to_give = totals["profit"] - totals["commission"] - ad_account - misc
    
    summary_data = [
        ["Summary", ""],
        ["Total Net Profit", format_rupees_full(totals['profit'])],
        [f"Platform Commission ({comm_pct}%)", format_rupees_full(totals['commission'])],
        [f"After Commission ({after_comm_pct}%)", format_rupees_full(totals['after_commission'])],
        ["Ad Account Charges", format_rupees_full(ad_account)],
        ["Miscellaneous Expenses", format_rupees_full(misc)],
        ["Funds to be Given", format_rupees_full(funds_to_give)]
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6AAF35')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('SPAN', (0, 0), (1, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E9')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    # Create filename with campaign name and month name (e.g., Medical_Emergency_Rahul_March_2026.pdf)
    safe_campaign_name = campaign['name'].replace(' ', '_').replace('-', '_')
    month_for_filename = format_month_name(report_month).replace(', ', '_').replace(' ', '_')
    filename = f"{safe_campaign_name}_{month_for_filename}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ EXCEL EXPORT ============
@api_router.get("/generate-excel/{campaign_id}/{report_month}")
async def generate_excel(campaign_id: str, report_month: str, user: dict = Depends(get_current_user)):
    campaign = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    settings = await get_settings()
    
    # Get entries for this month
    month_start = f"{report_month}-01"
    year, mon = int(report_month[:4]), int(report_month[5:7])
    if mon == 12:
        month_end = f"{year+1}-01-01"
    else:
        month_end = f"{year}-{mon+1:02d}-01"
    
    entries = await db.daily_entries.find({
        "campaign_id": campaign_id,
        "date": {"$gte": month_start, "$lt": month_end}
    }, {"_id": 0}).sort("date", 1).to_list(1000)
    
    settlement = await db.monthly_settlements.find_one({
        "campaign_id": campaign_id,
        "report_month": report_month
    }, {"_id": 0})
    
    month_display = format_month_name(report_month)
    comm_pct = campaign['commission_percentage']
    after_comm_pct = 100 - comm_pct
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    
    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='6AAF35', end_color='6AAF35', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='6AAF35')
    subtitle_font = Font(name='Calibri', size=11, color='444444')
    currency_format = '#,##0.00'
    total_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    total_font = Font(name='Calibri', bold=True, size=11)
    summary_header_fill = PatternFill(start_color='6AAF35', end_color='6AAF35', fill_type='solid')
    summary_highlight_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title section
    ws.merge_cells('A1:J1')
    ws['A1'] = 'Umeed Now Foundation - Monthly Report'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f'Campaign: {campaign["name"]}'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:J3')
    ws['A3'] = f'Month: {month_display} | Commission: {comm_pct}% | GST: {settings["gst_percentage"]}% | Gateway: {settings["gateway_percentage"]}%'
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Daily entries table header (row 5)
    headers = ['Date', 'Ad Spend', 'Ad+GST', 'Website', 'QR', 'Gateway', 'Revenue', 'Profit', f'Comm({comm_pct}%)', f'After({after_comm_pct}%)']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    totals = {"ad_spend": 0, "ad_spend_with_gst": 0, "website": 0, "qr": 0, "gateway": 0, "revenue": 0, "profit": 0, "commission": 0, "after_commission": 0}
    
    row = 6
    for entry in entries:
        after_comm = entry['net_profit'] - entry['platform_commission']
        date_obj = datetime.strptime(entry["date"], "%Y-%m-%d")
        
        ws.cell(row=row, column=1, value=date_obj.strftime("%d %b %Y")).border = thin_border
        ws.cell(row=row, column=2, value=round(entry['ad_spend'], 2)).border = thin_border
        ws.cell(row=row, column=3, value=round(entry['ad_spend_with_gst'], 2)).border = thin_border
        ws.cell(row=row, column=4, value=round(entry['website_collection'], 2)).border = thin_border
        ws.cell(row=row, column=5, value=round(entry['qr_collection'], 2)).border = thin_border
        ws.cell(row=row, column=6, value=round(entry['gateway_charge'], 2)).border = thin_border
        ws.cell(row=row, column=7, value=round(entry['total_revenue'], 2)).border = thin_border
        ws.cell(row=row, column=8, value=round(entry['net_profit'], 2)).border = thin_border
        ws.cell(row=row, column=9, value=round(entry['platform_commission'], 2)).border = thin_border
        ws.cell(row=row, column=10, value=round(after_comm, 2)).border = thin_border
        
        # Apply number format to currency columns
        for c in range(2, 11):
            ws.cell(row=row, column=c).number_format = currency_format
            ws.cell(row=row, column=c).alignment = Alignment(horizontal='right')
        
        totals["ad_spend"] += entry["ad_spend"]
        totals["ad_spend_with_gst"] += entry["ad_spend_with_gst"]
        totals["website"] += entry["website_collection"]
        totals["qr"] += entry["qr_collection"]
        totals["gateway"] += entry["gateway_charge"]
        totals["revenue"] += entry["total_revenue"]
        totals["profit"] += entry["net_profit"]
        totals["commission"] += entry["platform_commission"]
        totals["after_commission"] += after_comm
        row += 1
    
    # Totals row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL').font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = thin_border
    total_values = [totals['ad_spend'], totals['ad_spend_with_gst'], totals['website'], totals['qr'], totals['gateway'], totals['revenue'], totals['profit'], totals['commission'], totals['after_commission']]
    for c_idx, val in enumerate(total_values, 2):
        cell = ws.cell(row=total_row, column=c_idx, value=round(val, 2))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_format
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
    
    # Summary section (2 rows below totals)
    summary_start = total_row + 2
    ws.merge_cells(f'A{summary_start}:B{summary_start}')
    ws.cell(row=summary_start, column=1, value='Summary').font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    ws.cell(row=summary_start, column=1).fill = summary_header_fill
    ws.cell(row=summary_start, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=summary_start, column=2).fill = summary_header_fill
    
    ad_account = settlement.get("ad_account_charges", 0) if settlement else 0
    misc = settlement.get("miscellaneous_expenses", 0) if settlement else 0
    funds_to_give = totals["profit"] - totals["commission"] - ad_account - misc
    
    summary_items = [
        ("Total Net Profit", totals['profit']),
        (f"Platform Commission ({comm_pct}%)", totals['commission']),
        (f"After Commission ({after_comm_pct}%)", totals['after_commission']),
        ("Ad Account Charges", ad_account),
        ("Miscellaneous Expenses", misc),
        ("Funds to be Given", funds_to_give),
    ]
    
    for i, (label, value) in enumerate(summary_items):
        r = summary_start + 1 + i
        ws.cell(row=r, column=1, value=label).border = thin_border
        val_cell = ws.cell(row=r, column=2, value=round(value, 2))
        val_cell.number_format = currency_format
        val_cell.alignment = Alignment(horizontal='right')
        val_cell.border = thin_border
        # Highlight last row (Funds to be Given)
        if i == len(summary_items) - 1:
            ws.cell(row=r, column=1).font = total_font
            ws.cell(row=r, column=1).fill = summary_highlight_fill
            val_cell.font = total_font
            val_cell.fill = summary_highlight_fill
    
    # Column widths
    col_widths = [12, 12, 12, 14, 10, 12, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    safe_campaign_name = campaign['name'].replace(' ', '_').replace('-', '_')
    month_for_filename = format_month_name(report_month).replace(', ', '_').replace(' ', '_')
    filename = f"{safe_campaign_name}_{month_for_filename}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ STARTUP ============
@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.campaigns.create_index("id", unique=True)
    await db.daily_entries.create_index([("campaign_id", 1), ("date", 1)], unique=True)
    await db.monthly_settlements.create_index([("campaign_id", 1), ("report_month", 1)], unique=True)
    await db.login_attempts.create_index("identifier")
    
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@umeednow.org")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info("Admin password updated")
    
    # Initialize settings
    settings = await db.settings.find_one({"_id": "global"})
    if not settings:
        await db.settings.insert_one({"_id": "global", "gst_percentage": 18.0, "gateway_percentage": 2.6})
    
    # Seed demo data
    await seed_demo_data()
    
    # Write test credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write(f"# Test Credentials\n\n")
        f.write(f"## Admin Account\n")
        f.write(f"- Email: {admin_email}\n")
        f.write(f"- Password: {admin_password}\n")
        f.write(f"- Role: admin\n\n")
        f.write(f"## User Account\n")
        f.write(f"- Email: user@umeednow.org\n")
        f.write(f"- Password: user123\n")
        f.write(f"- Role: user\n\n")
        f.write(f"## Endpoints\n")
        f.write(f"- POST /api/auth/login\n")
        f.write(f"- POST /api/auth/logout\n")
        f.write(f"- GET /api/auth/me\n")

async def seed_demo_data():
    # Check if demo data exists
    campaign_count = await db.campaigns.count_documents({})
    if campaign_count > 0:
        return
    
    logger.info("Seeding demo data...")
    
    # Create demo user
    user_email = "user@umeednow.org"
    existing_user = await db.users.find_one({"email": user_email})
    if not existing_user:
        await db.users.insert_one({
            "email": user_email,
            "password_hash": hash_password("user123"),
            "name": "Demo User",
            "role": "user",
            "created_at": datetime.now(timezone.utc)
        })
    
    admin = await db.users.find_one({"role": "admin"})
    admin_id = str(admin["_id"]) if admin else "system"
    
    # Create demo campaigns
    campaigns = [
        {"name": "Medical Emergency - Rahul", "description": "Emergency medical treatment for Rahul", "commission_percentage": 25, "company_share": 4, "dev_share": 7, "himanshu_share": 7, "denim_share": 7},
        {"name": "Education Fund - Priya", "description": "College education fund for underprivileged student", "commission_percentage": 20, "company_share": 4, "dev_share": 6, "himanshu_share": 5, "denim_share": 5},
        {"name": "Flood Relief - Kerala", "description": "Disaster relief for flood victims", "commission_percentage": 15, "company_share": 3, "dev_share": 4, "himanshu_share": 4, "denim_share": 4}
    ]
    
    import random
    for camp_data in campaigns:
        campaign_id = str(ObjectId())
        await db.campaigns.insert_one({
            "id": campaign_id,
            **camp_data,
            "status": "active",
            "created_by": admin_id,
            "created_at": datetime.now(timezone.utc)
        })
        
        settings = await get_settings()
        
        # Add daily entries for last 60 days
        for days_ago in range(60):
            date = (datetime.now(timezone.utc) - timedelta(days=days_ago)).strftime("%Y-%m-%d")
            
            # Random data
            ad_spend = round(random.uniform(1000, 10000), 2)
            website_collection = round(random.uniform(5000, 50000), 2)
            qr_collection = round(random.uniform(500, 5000), 2)
            
            entry = {
                "ad_spend": ad_spend,
                "website_collection": website_collection,
                "qr_collection": qr_collection
            }
            calculated = calculate_daily_values(entry, settings, camp_data["commission_percentage"])
            
            entry_id = str(ObjectId())
            await db.daily_entries.insert_one({
                "id": entry_id,
                "campaign_id": campaign_id,
                "date": date,
                **entry,
                **calculated,
                "created_by": admin_id,
                "created_at": datetime.now(timezone.utc)
            })
        
        # Add monthly settlements for last 2 months
        for months_ago in [1, 2]:
            report_date = datetime.now(timezone.utc) - timedelta(days=30*months_ago)
            report_month = report_date.strftime("%Y-%m")
            
            await db.monthly_settlements.insert_one({
                "id": str(ObjectId()),
                "campaign_id": campaign_id,
                "report_month": report_month,
                "ad_account_charges": round(random.uniform(500, 2000), 2),
                "miscellaneous_expenses": round(random.uniform(100, 500), 2),
                "notes": "Monthly settlement",
                "entered_by": admin_id,
                "entered_on": datetime.now(timezone.utc)
            })
    
    logger.info("Demo data seeded successfully")

# Include the router
app.include_router(api_router)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000"), "https://fund-tracker-153.preview.emergentagent.com"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
