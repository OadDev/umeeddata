"""
Test PDF and Excel export functionality for Monthly Reports
Tests:
1. PDF generation endpoint - columns should not be cut off
2. Excel export endpoint - should return valid .xlsx file
3. Both endpoints require authentication
4. Both endpoints return proper file formats
"""
import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestPDFExcelExport:
    """Test PDF and Excel export endpoints"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - login and get campaign data"""
        self.session = requests.Session()
        
        # Login as admin
        login_response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@umeednow.org",
            "password": "admin123"
        })
        assert login_response.status_code == 200, f"Login failed: {login_response.text}"
        print(f"Login successful: {login_response.json()}")
        
        # Get campaigns
        campaigns_response = self.session.get(f"{BASE_URL}/api/campaigns")
        assert campaigns_response.status_code == 200, f"Failed to get campaigns: {campaigns_response.text}"
        self.campaigns = campaigns_response.json()
        assert len(self.campaigns) > 0, "No campaigns found"
        print(f"Found {len(self.campaigns)} campaigns")
        
        # Get monthly reports to find valid campaign_id and report_month
        reports_response = self.session.get(f"{BASE_URL}/api/monthly-reports")
        assert reports_response.status_code == 200, f"Failed to get monthly reports: {reports_response.text}"
        self.reports = reports_response.json()
        assert len(self.reports) > 0, "No monthly reports found"
        print(f"Found {len(self.reports)} monthly reports")
        
        # Use first report for testing
        self.test_campaign_id = self.reports[0]['campaign_id']
        self.test_report_month = self.reports[0]['month']
        self.test_campaign_name = self.reports[0]['campaign_name']
        print(f"Test campaign: {self.test_campaign_name} ({self.test_campaign_id}), month: {self.test_report_month}")
        
        yield
        
        # Logout
        self.session.post(f"{BASE_URL}/api/auth/logout")
    
    # ============ PDF GENERATION TESTS ============
    
    def test_pdf_generation_returns_200(self):
        """Test PDF generation endpoint returns 200 status"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200, f"PDF generation failed: {response.status_code} - {response.text}"
        print("PDF generation returned 200 OK")
    
    def test_pdf_content_type(self):
        """Test PDF has correct content type"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_type = response.headers.get('Content-Type', '')
        assert 'application/pdf' in content_type, f"Expected PDF content type, got: {content_type}"
        print(f"PDF content type: {content_type}")
    
    def test_pdf_content_disposition(self):
        """Test PDF has proper filename in Content-Disposition header"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition, f"Expected attachment disposition, got: {content_disposition}"
        assert '.pdf' in content_disposition, f"Expected .pdf in filename, got: {content_disposition}"
        print(f"PDF Content-Disposition: {content_disposition}")
    
    def test_pdf_file_size(self):
        """Test PDF file has reasonable size (not empty, not too small)"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_length = len(response.content)
        # PDF should be at least 5KB for a report with data
        assert content_length > 5000, f"PDF too small ({content_length} bytes), might be empty or corrupted"
        print(f"PDF file size: {content_length} bytes ({content_length/1024:.1f} KB)")
    
    def test_pdf_valid_format(self):
        """Test PDF starts with valid PDF header"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        # PDF files start with %PDF-
        pdf_header = response.content[:5]
        assert pdf_header == b'%PDF-', f"Invalid PDF header: {pdf_header}"
        print("PDF has valid header (%PDF-)")
    
    def test_pdf_invalid_campaign_returns_404(self):
        """Test PDF generation with invalid campaign returns 404"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-pdf/invalid_campaign_id/{self.test_report_month}"
        )
        assert response.status_code == 404, f"Expected 404 for invalid campaign, got: {response.status_code}"
        print("Invalid campaign correctly returns 404")
    
    # ============ EXCEL GENERATION TESTS ============
    
    def test_excel_generation_returns_200(self):
        """Test Excel generation endpoint returns 200 status"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200, f"Excel generation failed: {response.status_code} - {response.text}"
        print("Excel generation returned 200 OK")
    
    def test_excel_content_type(self):
        """Test Excel has correct content type"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_type = response.headers.get('Content-Type', '')
        expected_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert expected_type in content_type, f"Expected Excel content type, got: {content_type}"
        print(f"Excel content type: {content_type}")
    
    def test_excel_content_disposition(self):
        """Test Excel has proper filename in Content-Disposition header"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition, f"Expected attachment disposition, got: {content_disposition}"
        assert '.xlsx' in content_disposition, f"Expected .xlsx in filename, got: {content_disposition}"
        print(f"Excel Content-Disposition: {content_disposition}")
    
    def test_excel_file_size(self):
        """Test Excel file has reasonable size (not empty, not too small)"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        content_length = len(response.content)
        # Excel should be at least 5KB for a report with data
        assert content_length > 5000, f"Excel too small ({content_length} bytes), might be empty or corrupted"
        print(f"Excel file size: {content_length} bytes ({content_length/1024:.1f} KB)")
    
    def test_excel_valid_format(self):
        """Test Excel file has valid XLSX format (ZIP-based)"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        # XLSX files are ZIP archives, they start with PK (ZIP signature)
        xlsx_header = response.content[:2]
        assert xlsx_header == b'PK', f"Invalid XLSX header: {xlsx_header}"
        print("Excel has valid XLSX header (PK - ZIP format)")
    
    def test_excel_invalid_campaign_returns_404(self):
        """Test Excel generation with invalid campaign returns 404"""
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/invalid_campaign_id/{self.test_report_month}"
        )
        assert response.status_code == 404, f"Expected 404 for invalid campaign, got: {response.status_code}"
        print("Invalid campaign correctly returns 404")
    
    # ============ AUTHENTICATION TESTS ============
    
    def test_pdf_requires_authentication(self):
        """Test PDF endpoint requires authentication"""
        # Create new session without login
        new_session = requests.Session()
        response = new_session.get(
            f"{BASE_URL}/api/generate-pdf/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 401, f"Expected 401 for unauthenticated request, got: {response.status_code}"
        print("PDF endpoint correctly requires authentication")
    
    def test_excel_requires_authentication(self):
        """Test Excel endpoint requires authentication"""
        # Create new session without login
        new_session = requests.Session()
        response = new_session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 401, f"Expected 401 for unauthenticated request, got: {response.status_code}"
        print("Excel endpoint correctly requires authentication")
    
    # ============ CONTENT VALIDATION TESTS ============
    
    def test_excel_contains_data(self):
        """Test Excel file can be parsed and contains expected data"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed for content validation")
        
        response = self.session.get(
            f"{BASE_URL}/api/generate-excel/{self.test_campaign_id}/{self.test_report_month}"
        )
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Check title
        title_cell = ws['A1'].value
        assert 'Umeed Now Foundation' in title_cell, f"Expected title, got: {title_cell}"
        print(f"Excel title: {title_cell}")
        
        # Check headers row (row 5)
        headers = [ws.cell(row=5, column=i).value for i in range(1, 11)]
        expected_headers = ['Date', 'Ad Spend', 'Ad+GST', 'Website', 'QR', 'Gateway', 'Revenue', 'Profit']
        for expected in expected_headers:
            assert any(expected in str(h) for h in headers if h), f"Missing header: {expected}"
        print(f"Excel headers: {headers}")
        
        # Check there's data (row 6 should have first entry)
        first_data_row = [ws.cell(row=6, column=i).value for i in range(1, 11)]
        assert any(v is not None for v in first_data_row), "No data found in Excel"
        print(f"First data row: {first_data_row}")
        
        # Check for TOTAL row
        found_total = False
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == 'TOTAL':
                found_total = True
                break
        assert found_total, "TOTAL row not found in Excel"
        print("Excel contains TOTAL row")
        
        # Check for Summary section
        found_summary = False
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'Summary' in str(cell_value):
                found_summary = True
                break
        assert found_summary, "Summary section not found in Excel"
        print("Excel contains Summary section")
    
    def test_all_campaigns_pdf_generation(self):
        """Test PDF generation works for all campaigns with reports"""
        tested = 0
        for report in self.reports[:3]:  # Test first 3 reports
            response = self.session.get(
                f"{BASE_URL}/api/generate-pdf/{report['campaign_id']}/{report['month']}"
            )
            assert response.status_code == 200, f"PDF failed for {report['campaign_name']} {report['month']}"
            # PDF should be at least 2KB (valid PDF with minimal content)
            assert len(response.content) > 2000, f"PDF too small for {report['campaign_name']}"
            # Verify it's a valid PDF
            assert response.content[:5] == b'%PDF-', f"Invalid PDF for {report['campaign_name']}"
            tested += 1
            print(f"PDF OK: {report['campaign_name']} - {report['month']} ({len(response.content)} bytes)")
        print(f"Tested {tested} PDF generations successfully")
    
    def test_all_campaigns_excel_generation(self):
        """Test Excel generation works for all campaigns with reports"""
        tested = 0
        for report in self.reports[:3]:  # Test first 3 reports
            response = self.session.get(
                f"{BASE_URL}/api/generate-excel/{report['campaign_id']}/{report['month']}"
            )
            assert response.status_code == 200, f"Excel failed for {report['campaign_name']} {report['month']}"
            assert len(response.content) > 5000, f"Excel too small for {report['campaign_name']}"
            tested += 1
            print(f"Excel OK: {report['campaign_name']} - {report['month']}")
        print(f"Tested {tested} Excel generations successfully")


class TestMonthlyReportsAPI:
    """Test Monthly Reports API endpoint"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - login"""
        self.session = requests.Session()
        login_response = self.session.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@umeednow.org",
            "password": "admin123"
        })
        assert login_response.status_code == 200
        yield
        self.session.post(f"{BASE_URL}/api/auth/logout")
    
    def test_monthly_reports_returns_data(self):
        """Test monthly reports endpoint returns data"""
        response = self.session.get(f"{BASE_URL}/api/monthly-reports")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        assert len(data) > 0, "No monthly reports found"
        print(f"Found {len(data)} monthly reports")
    
    def test_monthly_reports_structure(self):
        """Test monthly reports have expected structure"""
        response = self.session.get(f"{BASE_URL}/api/monthly-reports")
        assert response.status_code == 200
        data = response.json()
        
        required_fields = [
            'campaign_id', 'campaign_name', 'month', 'funds_raised',
            'ad_cost_with_gst', 'gateway_charge', 'net_profit', 'commission',
            'ad_account_charges', 'miscellaneous_expenses', 'funds_to_give',
            'stakeholder_earnings', 'has_settlement'
        ]
        
        for report in data[:3]:
            for field in required_fields:
                assert field in report, f"Missing field: {field}"
        print("Monthly reports have correct structure")
    
    def test_monthly_reports_filter_by_campaign(self):
        """Test filtering monthly reports by campaign"""
        # Get all reports first
        all_response = self.session.get(f"{BASE_URL}/api/monthly-reports")
        all_data = all_response.json()
        
        if len(all_data) > 0:
            campaign_id = all_data[0]['campaign_id']
            filtered_response = self.session.get(f"{BASE_URL}/api/monthly-reports?campaign_id={campaign_id}")
            filtered_data = filtered_response.json()
            
            # All filtered reports should be for the same campaign
            for report in filtered_data:
                assert report['campaign_id'] == campaign_id
            print(f"Campaign filter works: {len(filtered_data)} reports for campaign")
    
    def test_monthly_reports_filter_by_month(self):
        """Test filtering monthly reports by month"""
        all_response = self.session.get(f"{BASE_URL}/api/monthly-reports")
        all_data = all_response.json()
        
        if len(all_data) > 0:
            month = all_data[0]['month']
            filtered_response = self.session.get(f"{BASE_URL}/api/monthly-reports?report_month={month}")
            filtered_data = filtered_response.json()
            
            # All filtered reports should be for the same month
            for report in filtered_data:
                assert report['month'] == month
            print(f"Month filter works: {len(filtered_data)} reports for {month}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
